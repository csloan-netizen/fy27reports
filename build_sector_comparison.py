import openpyxl
import html as html_module
import os
from collections import defaultdict

BASE = os.path.dirname(os.path.abspath(__file__))

wb = openpyxl.load_workbook(
    os.path.join(BASE, 'data', 'ap_pipegen.xlsx'),
    read_only=True, data_only=True
)

SECTOR_ORDER = ['AMER TMT/CBS', 'AMER PACE', 'AMER REG', 'AMER Pub Sec', 'LATAM']
SECTOR_COLORS = {
    'AMER TMT/CBS': '#0070d2',
    'AMER PACE':    '#00a1e0',
    'AMER REG':     '#1b7a3e',
    'AMER Pub Sec': '#6b3fa0',
    'LATAM':        '#c23934',
}

# ── Load Results: grand total bookings + pipegen per AP per sector ─────────────
ws_res = wb['Results']
res_rows = list(ws_res.iter_rows(values_only=True))

sectors = {s: {'bookings': 0, 'pipegen': 0, 'aps': 0,
               'q1_bookings': 0, 'q1_pipegen': 0,
               'q2_bookings': 0, 'q2_pipegen': 0,
               'q3_bookings': 0, 'q3_pipegen': 0,
               'q3_open_pipe': 0}
           for s in SECTOR_ORDER}

for r in res_rows[2:]:
    if not r[1] or not r[3] or str(r[1]) == 'AMERS':
        continue
    ou = str(r[1])
    if ou not in sectors:
        continue
    s = sectors[ou]
    s['aps']        += 1
    s['q1_bookings'] += float(r[4])  if r[4]  is not None else 0
    s['q1_pipegen']  += float(r[5])  if r[5]  is not None else 0
    s['q2_bookings'] += float(r[6])  if r[6]  is not None else 0
    s['q2_pipegen']  += float(r[7])  if r[7]  is not None else 0
    s['q3_bookings'] += float(r[8])  if r[8]  is not None else 0
    s['q3_pipegen']  += float(r[9])  if r[9]  is not None else 0
    s['bookings']    += float(r[10]) if r[10] is not None else 0
    s['pipegen']     += float(r[11]) if r[11] is not None else 0

# ── Load Export: Q3 open pipe ─────────────────────────────────────────────────
ws_exp = wb['Export']
exp_rows = list(ws_exp.iter_rows(values_only=True))
for r in exp_rows[1:]:
    if r[0] != 'Q3':
        continue
    ou = str(r[7]) if r[7] else ''
    if ou not in sectors:
        continue
    sectors[ou]['q3_open_pipe'] += float(r[9]) if r[9] is not None else 0

# ── Derived metrics ───────────────────────────────────────────────────────────
for s in sectors.values():
    s['book_per_ap']  = s['bookings'] / s['aps']  if s['aps'] else 0
    s['pipe_per_ap']  = s['pipegen']  / s['aps']  if s['aps'] else 0
    s['book_pct']     = s['bookings'] / s['pipegen'] * 100 if s['pipegen'] else 0

# ── Totals ────────────────────────────────────────────────────────────────────
total = {k: sum(sectors[s][k] for s in SECTOR_ORDER)
         for k in ['bookings','pipegen','q1_bookings','q1_pipegen',
                   'q2_bookings','q2_pipegen','q3_bookings','q3_open_pipe',
                   'q3_pipegen','aps']}
total['book_per_ap'] = total['bookings'] / total['aps'] if total['aps'] else 0
total['pipe_per_ap'] = total['pipegen']  / total['aps'] if total['aps'] else 0
total['book_pct']    = total['bookings'] / total['pipegen'] * 100 if total['pipegen'] else 0

# ── Helpers ───────────────────────────────────────────────────────────────────
def h(s):
    return html_module.escape(str(s) if s else '')

def fmt(v, zero_dash=True):
    if v == 0 and zero_dash:
        return '<span class="zero">—</span>'
    return '${:,.0f}'.format(v)

def fmt_pct(v):
    return '{:.1f}%'.format(v)

def bar(val, max_val, color, height=20):
    if max_val == 0:
        return ''
    pct = min(100, val / max_val * 100)
    return '<div class="hbar" style="width:{:.1f}%;background:{};height:{}px"></div>'.format(pct, color, height)

max_pipegen  = max(sectors[s]['pipegen']     for s in SECTOR_ORDER)
max_bookings = max(sectors[s]['bookings']    for s in SECTOR_ORDER)
max_per_ap   = max(sectors[s]['pipe_per_ap'] for s in SECTOR_ORDER)

# ── Table rows ────────────────────────────────────────────────────────────────
def make_row(name, d, is_total=False):
    color = SECTOR_COLORS.get(name, '#666')
    highlight = 'tmt-row' if name == 'AMER TMT/CBS' else ''
    if is_total:
        highlight = 'total-row'
        color = '#444'
    return '''<tr class="data-row {hl}">
  <td class="name"><span class="dot" style="background:{c}"></span>{nm}</td>
  <td class="num">{aps}</td>
  <td class="money">{q1b}</td>
  <td class="money pipegen">{q1p}</td>
  <td class="money">{q2b}</td>
  <td class="money pipegen">{q2p}</td>
  <td class="money">{q3b}</td>
  <td class="money open">{q3op}</td>
  <td class="money pipegen">{q3p}</td>
  <td class="money tot-book">{tb}</td>
  <td class="money tot-pipe">{tp}
    <div class="bar-cell">{pipebar}</div>
  </td>
  <td class="money">{bpa}</td>
  <td class="money">{ppa}
    <div class="bar-cell">{ppabar}</div>
  </td>
  <td class="pct {bpct_cls}">{bpct}</td>
</tr>'''.format(
        hl=highlight, c=h(color), nm=h(name),
        aps=d['aps'],
        q1b=fmt(d['q1_bookings']), q1p=fmt(d['q1_pipegen']),
        q2b=fmt(d['q2_bookings']), q2p=fmt(d['q2_pipegen']),
        q3b=fmt(d['q3_bookings']), q3op=fmt(d['q3_open_pipe']),
        q3p=fmt(d['q3_pipegen']),
        tb=fmt(d['bookings']), tp=fmt(d['pipegen']),
        pipebar='' if is_total else bar(d['pipegen'], max_pipegen, color),
        bpa=fmt(d['book_per_ap']), ppa=fmt(d['pipe_per_ap']),
        ppabar='' if is_total else bar(d['pipe_per_ap'], max_per_ap, color),
        bpct=fmt_pct(d['book_pct']),
        bpct_cls='hi' if d['book_pct'] >= total['book_pct'] else 'lo',
    )

rows_html = [make_row(s, sectors[s]) for s in SECTOR_ORDER]
rows_html.append(make_row('All Sectors', total, is_total=True))
rows_str = '\n'.join(rows_html)

# ── Summary bar chart data (for sparkline-style bars at top) ─────────────────
def pct_of_total(sector, field):
    t = total[field]
    return sectors[sector][field] / t * 100 if t else 0

summary_cards = []
for s in SECTOR_ORDER:
    d = sectors[s]
    color = SECTOR_COLORS[s]
    summary_cards.append('''<div class="scard {cls}">
  <div class="scard-name"><span class="dot" style="background:{c}"></span>{nm}</div>
  <div class="scard-stats">
    <div class="scard-stat"><div class="scard-val">{aps}</div><div class="scard-lbl">APs</div></div>
    <div class="scard-stat"><div class="scard-val">{pipe}</div><div class="scard-lbl">FY27 Pipegen</div></div>
    <div class="scard-stat"><div class="scard-val">{book}</div><div class="scard-lbl">FY27 Bookings</div></div>
    <div class="scard-stat"><div class="scard-val">{q3op}</div><div class="scard-lbl">Q3 Open Pipe</div></div>
    <div class="scard-stat"><div class="scard-val">{ppa}</div><div class="scard-lbl">Pipegen / AP</div></div>
    <div class="scard-stat"><div class="scard-val">{bpct}</div><div class="scard-lbl">Book / Pipe %</div></div>
  </div>
  <div class="scard-share">
    <div class="share-lbl">Share of total pipegen</div>
    <div class="share-bar-wrap"><div class="share-bar" style="width:{share:.1f}%;background:{c}"></div></div>
    <div class="share-val">{share:.1f}%</div>
  </div>
</div>'''.format(
        cls='tmt-card' if s == 'AMER TMT/CBS' else '',
        c=h(color), nm=h(s),
        aps=d['aps'],
        pipe='${:,.0f}'.format(d['pipegen']),
        book='${:,.0f}'.format(d['bookings']),
        q3op='${:,.0f}'.format(d['q3_open_pipe']),
        ppa='${:,.0f}'.format(d['pipe_per_ap']),
        bpct=fmt_pct(d['book_pct']),
        share=pct_of_total(s, 'pipegen'),
    ))

cards_str = '\n'.join(summary_cards)

html_doc = """<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<title>AMERS AP Sector Comparison FY27</title>
<style>
* {{ box-sizing: border-box; margin: 0; padding: 0; }}
body {{ font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Arial, sans-serif; font-size: 12px; background: #f0f4f8; color: #1a1a2e; }}
.header {{ padding: 20px 24px 8px; }}
h1 {{ font-size: 20px; color: #032D60; margin-bottom: 4px; }}
.subtitle {{ color: #555; font-size: 12px; margin-bottom: 18px; }}

/* Summary cards */
.cards {{ display: flex; gap: 12px; padding: 0 24px 22px; flex-wrap: wrap; }}
.scard {{ background: white; border-radius: 8px; padding: 14px 16px; box-shadow: 0 1px 4px rgba(0,0,0,.1); flex: 1; min-width: 180px; border-top: 3px solid #ccc; }}
.tmt-card {{ border-top-color: #0070d2; box-shadow: 0 2px 10px rgba(0,112,210,.2); }}
.scard-name {{ font-weight: 700; font-size: 12px; color: #032D60; margin-bottom: 10px; display: flex; align-items: center; gap: 6px; }}
.scard-stats {{ display: grid; grid-template-columns: 1fr 1fr; gap: 8px 12px; margin-bottom: 10px; }}
.scard-stat .scard-val {{ font-size: 13px; font-weight: 700; color: #032D60; }}
.scard-stat .scard-lbl {{ font-size: 9px; color: #888; text-transform: uppercase; letter-spacing: .4px; margin-top: 1px; }}
.share-lbl {{ font-size: 9px; color: #aaa; margin-bottom: 3px; text-transform: uppercase; letter-spacing: .4px; }}
.share-bar-wrap {{ background: #e8ecf2; border-radius: 3px; height: 6px; width: 100%; margin-bottom: 3px; }}
.share-bar {{ height: 6px; border-radius: 3px; }}
.share-val {{ font-size: 10px; color: #666; }}

/* Dot */
.dot {{ width: 10px; height: 10px; border-radius: 50%; display: inline-block; flex-shrink: 0; }}

/* Table */
.wrapper {{ padding: 0 16px 40px; overflow-x: auto; }}
table {{ border-collapse: collapse; width: 100%; background: white; box-shadow: 0 1px 6px rgba(0,0,0,.1); border-radius: 8px; overflow: hidden; min-width: 1060px; }}
thead tr.qhdr {{ background: #032D60; color: white; }}
thead tr.qhdr th {{ padding: 8px 10px; font-size: 10px; font-weight: 600; letter-spacing: .5px; text-align: center; white-space: nowrap; }}
thead tr.qhdr th.q2hdr {{ background: #0b4f8a; }}
thead tr.qhdr th.q3hdr {{ background: #0070d2; }}
thead tr.qhdr th.tothdr {{ background: #00396b; }}
thead tr.qhdr th.effhdr {{ background: #1b5e20; }}
thead tr.qhdr th.namehdr {{ text-align: left; }}
thead tr.fhdr {{ background: #1a3a5c; color: #cde; }}
thead tr.fhdr th {{ padding: 7px 10px; font-size: 10px; font-weight: 500; letter-spacing: .3px; white-space: nowrap; }}
thead tr.fhdr th.pipegen {{ color: #88ccff; }}
thead tr.fhdr th.open {{ color: #aaddff; }}
thead tr.fhdr th.eff {{ color: #a8d5a2; }}
td {{ padding: 8px 10px; vertical-align: middle; border-bottom: 1px solid #eef0f3; }}
tr.tmt-row td {{ background: #f0f7ff !important; }}
tr.tmt-row td.name {{ border-left: 3px solid #0070d2; }}
tr.total-row td {{ background: #032D60 !important; color: white; font-weight: 700; border-top: 2px solid #0070d2; }}
tr.total-row td.money, tr.total-row td.pct, tr.total-row td.num {{ color: #cce5ff; }}
tr.data-row:not(.tmt-row):not(.total-row):hover td {{ background: #f8fafc; }}
td.name {{ font-weight: 600; color: #032D60; width: 150px; border-right: 2px solid #dde3ed; white-space: nowrap; display: flex; align-items: center; gap: 6px; }}
tr.total-row td.name {{ color: white; }}
td.money {{ text-align: right; white-space: nowrap; font-variant-numeric: tabular-nums; width: 95px; }}
td.num {{ text-align: center; width: 45px; color: #555; }}
td.money .zero {{ color: #ccc; }}
td.pipegen {{ background: #f5faff; }}
td.open {{ background: #edf5ff; }}
td.tot-book {{ background: #f8f5ff; font-weight: 600; }}
td.tot-pipe {{ background: #eef3ff; font-weight: 700; color: #032D60; width: 120px; }}
tr.total-row td.tot-pipe {{ color: #88ccff; }}
td.pct {{ text-align: center; width: 65px; font-weight: 600; border-left: 1px solid #dde3ed; }}
td.pct.hi {{ color: #1b7a3e; }}
td.pct.lo {{ color: #c23934; }}
tr.total-row td.pct {{ color: #a0d8b0; }}
.bar-cell {{ margin-top: 4px; }}
.hbar {{ border-radius: 2px; }}
</style>
</head>
<body>
<div class="header">
  <h1>AMERS — AP Sector Comparison FY27</h1>
  <p class="subtitle">Bookings &amp; pipeline generation by sector &nbsp;·&nbsp; FY27 Q1–Q3 &nbsp;·&nbsp; as of 2026-08-06</p>
</div>

<div class="cards">
{cards}
</div>

<div class="wrapper">
<table>
<thead>
<tr class="qhdr">
  <th class="namehdr" rowspan="2">Sector</th>
  <th rowspan="2" style="text-align:center">APs</th>
  <th colspan="2">Q1</th>
  <th class="q2hdr" colspan="2">Q2</th>
  <th class="q3hdr" colspan="3">Q3</th>
  <th class="tothdr" colspan="2">Grand Total</th>
  <th class="effhdr" colspan="3">Efficiency</th>
</tr>
<tr class="fhdr">
  <th>Bookings</th><th class="pipegen">Pipegen</th>
  <th>Bookings</th><th class="pipegen">Pipegen</th>
  <th>Bookings</th><th class="open">Open Pipe</th><th class="pipegen">Pipegen</th>
  <th>Bookings</th><th class="pipegen">Pipegen</th>
  <th class="eff">Book / AP</th><th class="eff">Pipe / AP</th><th class="eff">Book %</th>
</tr>
</thead>
<tbody>
{rows}
</tbody>
</table>
</div>
</body>
</html>""".format(cards=cards_str, rows=rows_str)

out_path = os.path.join(BASE, 'tmt_sector_comparison.html')
with open(out_path, 'w', encoding='utf-8') as f:
    f.write(html_doc)

print("Written to {}".format(out_path))
for s in SECTOR_ORDER:
    d = sectors[s]
    print("  {:20s}  aps={:2d}  pipegen={:>15s}  bookings={:>14s}  book%={:.1f}%".format(
        s, d['aps'],
        '${:,.0f}'.format(d['pipegen']),
        '${:,.0f}'.format(d['bookings']),
        d['book_pct']
    ))
