import csv
import re
import html
from collections import defaultdict

import os
BASE = os.path.dirname(os.path.abspath(__file__))

with open(os.path.join(BASE, 'data', 'license.csv'), 'r', encoding='latin-1') as f:
    license_rows = list(csv.DictReader(f))

with open(os.path.join(BASE, 'data', 'services.csv'), 'r', encoding='latin-1') as f:
    services_rows = list(csv.DictReader(f))

import openpyxl
wb = openpyxl.load_workbook(
    os.path.join(BASE, 'data', 'account_mapping.xlsx'),
    read_only=True, data_only=True
)
map_rows = []
for sheet in wb.sheetnames:
    for row in wb[sheet].iter_rows(values_only=True):
        map_rows.append([str(v).strip() if v is not None else '' for v in row])

# ── Partner lookup ────────────────────────────────────────────────────────────
def _norm_simple(s):
    return re.sub(r'[^a-z0-9 ]', '', s.lower()).strip()

partner_by_acct = {}
for r in map_rows:
    if len(r) < 15 or not r[0].strip():
        continue
    partner = r[0].strip()
    for val in [r[7].strip(), r[14].strip()]:
        if val:
            partner_by_acct[_norm_simple(val)] = partner

PARTNER_OVERRIDES = {
    _norm_simple('Datasite LLC'):      'Alex Lee',
    _norm_simple('Discovery, Inc.'):   'John-Michael Knowles',
    _norm_simple('Ul India Pvt Ltd'):  'Joseph Carpenter',
}

def get_partner(acct_name):
    n = _norm_simple(acct_name)
    if n in PARTNER_OVERRIDES:
        return PARTNER_OVERRIDES[n]
    return partner_by_acct.get(n, '')

# ── Helpers ───────────────────────────────────────────────────────────────────
def norm(s):
    s = s.lower().strip()
    for suf in [', inc.', ', inc', ' inc.', ' inc', ' llc', ', llc', ', ltd.', ', ltd',
                ' ltd.', ', corp.', ', corp', ' corp.', ', co.', ', co', ' co.',
                ', plc', ' plc', ', ag', ' ag', ', nv', ', bv', ', sa', ', s.a.',
                ' technologies', ' technology', ' solutions', ' systems',
                ' group', ' holdings', ', limited']:
        while s.endswith(suf):
            s = s[:-len(suf)].rstrip(', ')
    return re.sub(r'\s+', ' ', s).strip()

def fmt_money(s):
    try:
        v = float(str(s).replace(',', ''))
        return '${:,.0f}'.format(v) if v else '-'
    except:
        return s or '-'

def h(s):
    return html.escape(str(s) if s else '')

# ── Services lookups ──────────────────────────────────────────────────────────
svc_by_lic_oppname = defaultdict(list)
for r in services_rows:
    rel = r.get('Related License Oppty', '').strip()
    if rel:
        svc_by_lic_oppname[rel].append(r)

def extract_acct_prefix(svc_opp_name):
    for sw in ['ARI Investment', 'MS -', 'DF-', 'TEST ']:
        if svc_opp_name.startswith(sw):
            return None
    return svc_opp_name.split(' - ')[0].strip()

svc_by_norm_acct = defaultdict(list)
for r in services_rows:
    prefix = extract_acct_prefix(r['Opportunity Name'].strip())
    if prefix:
        svc_by_norm_acct[norm(prefix)].append(r)

svc_owner_by_norm_acct = {}
for n, svcs in svc_by_norm_acct.items():
    owners = [s.get('Opportunity Owner', '').strip() for s in svcs if s.get('Opportunity Owner', '').strip()]
    if owners:
        svc_owner_by_norm_acct[n] = owners[0]

svc_owner_by_lic_oppname = {}
for opp_name, svcs in svc_by_lic_oppname.items():
    for s in svcs:
        owner = s.get('Opportunity Owner', '').strip()
        if owner:
            svc_owner_by_lic_oppname[opp_name] = owner
            break

def get_svc_owner(lic_opp_name, acct):
    if lic_opp_name in svc_owner_by_lic_oppname:
        return svc_owner_by_lic_oppname[lic_opp_name]
    if norm(acct) in svc_owner_by_norm_acct:
        return svc_owner_by_norm_acct[norm(acct)]
    return get_partner(acct)

# ── Group & sort license opps ─────────────────────────────────────────────────
lic_by_acct = defaultdict(list)
for r in license_rows:
    acct = r['Account Name: Future Combo Company'].strip()
    lic_by_acct[acct].append(r)

def total_amt(opps):
    t = 0
    for o in opps:
        try:
            t += float(o.get('Amount (converted)', '0').replace(',', '') or 0)
        except:
            pass
    return t

sorted_accounts = sorted(lic_by_acct.items(), key=lambda x: total_amt(x[1]), reverse=True)
for _, opps in sorted_accounts:
    opps.sort(key=lambda r: float(r.get('Amount (converted)', '0').replace(',', '') or 0), reverse=True)

# ── Build tbody groups ────────────────────────────────────────────────────────
tbodies_html = []
all_partners = set()

for acct, opps in sorted_accounts:
    acct_total = total_amt(opps)
    partner = get_partner(acct)
    if partner:
        all_partners.add(partner)
    first_acct_row = True

    acct_norm = norm(acct)
    acct_level_svcs = svc_by_norm_acct.get(acct_norm, [])

    rows_html = []

    for i, lic in enumerate(opps):
        lic_name = lic['Opportunity Name'].strip()
        lic_amt = fmt_money(lic.get('Amount (converted)', ''))
        lic_ae = lic.get('Opportunity Owner: Full Name', '').strip()
        lic_close = lic.get('Close Date', '').strip()

        direct_svcs = svc_by_lic_oppname.get(lic_name, [])

        if not direct_svcs and first_acct_row and acct_level_svcs:
            svcs_to_show = acct_level_svcs
        elif direct_svcs:
            svcs_to_show = direct_svcs
        else:
            svcs_to_show = []

        acct_cell_val = ''
        if first_acct_row:
            acct_cell_val = '<strong>{}</strong><span class="total">{}</span>'.format(
                h(acct), fmt_money(str(acct_total))
            )

        if svcs_to_show:
            for j, svc in enumerate(svcs_to_show):
                svc_name = svc['Opportunity Name'].strip()
                svc_raw = svc.get('Amount (converted)', '0')
                try:
                    svc_val = float(svc_raw.replace(',', ''))
                except:
                    svc_val = 0
                if svc_val == 0:
                    inv_raw = svc.get('Investment Amount', '0')
                    try:
                        inv_val = float(inv_raw.replace(',', ''))
                    except:
                        inv_val = 0
                    svc_amt = fmt_money(inv_raw) if inv_val else '-'
                else:
                    svc_amt = fmt_money(svc_raw)
                svc_owner = svc.get('Opportunity Owner', '').strip()
                svc_close = svc.get('Close Date', '').strip()

                row_class = 'svc-extra' if j > 0 else ''
                if j == 0:
                    lic_cells = '''
  <td class="acct">{}</td>
  <td class="money">{}</td>
  <td class="ae">{}</td>
  <td class="opp-name">{}</td>
  <td class="date">{}</td>'''.format(acct_cell_val, h(lic_amt), h(lic_ae), h(lic_name), h(lic_close))
                else:
                    lic_cells = '''
  <td class="acct"></td>
  <td class="money"></td>
  <td class="ae"></td>
  <td class="opp-name"></td>
  <td class="date"></td>'''

                rows_html.append('''<tr class="data-row has-svc {rc}">{lc}
  <td class="svc-col opp-name">{sn}</td>
  <td class="svc-col money">{sa}</td>
  <td class="svc-col ae">{so}</td>
  <td class="svc-col date">{sc}</td>
</tr>'''.format(rc=row_class, lc=lic_cells, sn=h(svc_name), sa=svc_amt, so=h(svc_owner), sc=h(svc_close)))

        else:
            acct_svc_owner = get_svc_owner(lic_name, acct)
            no_svc_flag = ' no-svc' if not svcs_to_show else ''
            rows_html.append('''<tr class="data-row{f}">
  <td class="acct">{av}</td>
  <td class="money">{am}</td>
  <td class="ae">{ae}</td>
  <td class="opp-name">{on}</td>
  <td class="date">{cd}</td>
  <td class="svc-col no-attach" colspan="2">&#9888; No services attach</td>
  <td class="svc-col ae">{so}</td>
  <td class="svc-col date"></td>
</tr>'''.format(f=no_svc_flag, av=acct_cell_val, am=h(lic_amt), ae=h(lic_ae),
                on=h(lic_name), cd=h(lic_close), so=h(acct_svc_owner)))

        if first_acct_row:
            first_acct_row = False

    rows_html.append('<tr class="divider"><td colspan="9"></td></tr>')

    tbodies_html.append(
        '<tbody class="acct-group" data-account="{acct}" data-partner="{pt}" data-total="{tot}">\n{rows}\n</tbody>'.format(
            acct=h(acct), pt=h(partner), tot=int(acct_total), rows='\n'.join(rows_html)
        )
    )

rows_str = '\n'.join(tbodies_html)

# ── Filter/sort dropdowns ─────────────────────────────────────────────────────
acct_options = '\n'.join(
    '<option value="{v}">{v}</option>'.format(v=h(a))
    for a, _ in sorted(sorted_accounts, key=lambda x: x[0].lower())
)
partner_options = '\n'.join(
    '<option value="{v}">{v}</option>'.format(v=h(p))
    for p in sorted(all_partners)
)

# ── Stats ─────────────────────────────────────────────────────────────────────
total_pipeline = sum(total_amt(v) for _, v in sorted_accounts)
accts_with_svc = sum(
    1 for acct, opps in sorted_accounts
    if any(
        svc_by_lic_oppname.get(o['Opportunity Name'].strip()) or svc_by_norm_acct.get(norm(acct))
        for o in opps
    )
)

# ── HTML ──────────────────────────────────────────────────────────────────────
html_doc = """<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<title>AMER TMT - License + Services Q3</title>
<style>
* {{ box-sizing: border-box; margin: 0; padding: 0; }}
body {{ font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Arial, sans-serif; font-size: 12px; background: #f0f4f8; color: #1a1a2e; }}
.header {{ padding: 20px 24px 0; }}
h1 {{ font-size: 20px; color: #032D60; margin-bottom: 4px; }}
.subtitle {{ color: #555; font-size: 12px; margin-bottom: 14px; }}
.stats {{ display: flex; gap: 12px; padding: 0 24px 14px; flex-wrap: wrap; }}
.stat {{ background: white; border-radius: 6px; padding: 10px 18px; box-shadow: 0 1px 3px rgba(0,0,0,.1); min-width: 130px; }}
.stat .val {{ font-size: 22px; font-weight: 700; color: #032D60; }}
.stat .lbl {{ font-size: 10px; color: #666; margin-top: 2px; text-transform: uppercase; letter-spacing: .5px; }}
.controls {{ display: flex; gap: 12px; padding: 0 24px 14px; flex-wrap: wrap; align-items: flex-end; }}
.control-group {{ display: flex; flex-direction: column; gap: 4px; }}
.control-group label {{ font-size: 10px; font-weight: 600; color: #444; text-transform: uppercase; letter-spacing: .4px; }}
.control-group select, .control-group input {{ font-size: 12px; border: 1px solid #ccd3de; border-radius: 4px; padding: 5px 8px; background: white; color: #1a1a2e; min-width: 200px; }}
.control-group select:focus, .control-group input:focus {{ outline: none; border-color: #0070d2; box-shadow: 0 0 0 2px rgba(0,112,210,.2); }}
.sort-group {{ display: flex; gap: 6px; }}
.sort-btn {{ font-size: 11px; border: 1px solid #ccd3de; border-radius: 4px; padding: 5px 10px; background: white; color: #333; cursor: pointer; white-space: nowrap; }}
.sort-btn:hover {{ background: #f0f4f8; border-color: #0070d2; color: #0070d2; }}
.sort-btn.active {{ background: #032D60; border-color: #032D60; color: white; }}
.sort-btn.active.asc::after {{ content: ' ↑'; }}
.sort-btn.active.desc::after {{ content: ' ↓'; }}
.legend {{ display: flex; gap: 20px; padding: 0 24px 12px; font-size: 11px; align-items: center; }}
.legend-item {{ display: flex; align-items: center; gap: 6px; }}
.dot {{ width: 12px; height: 12px; border-radius: 2px; flex-shrink: 0; }}
.dot-no {{ background: #fff3cd; border: 1px solid #ffc107; }}
.dot-yes {{ background: #e6f4ea; border: 1px solid #34a853; }}
.result-count {{ padding: 0 24px 8px; font-size: 11px; color: #666; }}
.wrapper {{ padding: 0 16px 40px; overflow-x: auto; }}
table {{ border-collapse: collapse; width: 100%; background: white; box-shadow: 0 1px 6px rgba(0,0,0,.1); border-radius: 8px; overflow: hidden; min-width: 960px; }}
thead tr {{ background: #032D60; color: white; }}
thead th {{ padding: 10px 10px; text-align: left; font-size: 11px; font-weight: 600; letter-spacing: .4px; white-space: nowrap; }}
thead .svc-hdr {{ background: #0070d2; }}
td {{ padding: 6px 10px; vertical-align: top; border-bottom: 1px solid #eef0f3; line-height: 1.45; }}
tbody.acct-group.hidden {{ display: none; }}
tr.divider td {{ background: #e8ecf2 !important; height: 3px; padding: 0; border: none; }}
td.acct {{ font-weight: 600; color: #032D60; width: 170px; border-right: 2px solid #dde3ed; }}
td.acct .total {{ font-size: 10px; font-weight: 400; color: #888; display: block; margin-top: 3px; }}
td.money {{ text-align: right; white-space: nowrap; font-variant-numeric: tabular-nums; width: 90px; }}
td.ae {{ width: 110px; white-space: nowrap; }}
td.date {{ white-space: nowrap; width: 90px; }}
td.opp-name {{ max-width: 200px; }}
td.svc-col {{ background: #f5f9ff; }}
td.no-attach {{ background: #fff8e1; color: #856404; font-style: italic; font-size: 11px; border-left: 3px solid #ffc107; }}
tr.no-svc td:not(.svc-col):not(.no-attach) {{ background: #fffdf4; }}
tr.has-svc td:not(.svc-col) {{ background: #f9fffb; }}
tr.svc-extra td {{ border-top: none; }}
tr.svc-extra td.svc-col {{ border-top: 1px dashed #cce0ff; }}
</style>
</head>
<body>
<div class="header">
  <h1>AMER TMT — License &amp; Services Opportunities</h1>
  <p class="subtitle">Q3 &nbsp;·&nbsp; {acct_count:,} accounts &nbsp;·&nbsp; {lic_count:,} license opps &nbsp;·&nbsp; {svc_count:,} services opps &nbsp;·&nbsp; sorted by total license value descending</p>
</div>
<div class="stats">
  <div class="stat"><div class="val">{acct_count:,}</div><div class="lbl">Accounts</div></div>
  <div class="stat"><div class="val">{lic_count:,}</div><div class="lbl">License Opps</div></div>
  <div class="stat"><div class="val">{svc_count:,}</div><div class="lbl">Services Opps</div></div>
  <div class="stat"><div class="val">${pipeline}</div><div class="lbl">Total License Pipeline</div></div>
  <div class="stat"><div class="val">{accts_svc}</div><div class="lbl">Accts w/ Services</div></div>
  <div class="stat"><div class="val">{accts_no_svc}</div><div class="lbl">Accts w/o Services</div></div>
</div>

<div class="controls">
  <div class="control-group">
    <label>Filter by Account</label>
    <select id="filterAccount" onchange="applyFilters()">
      <option value="">All accounts</option>
      {acct_options}
    </select>
  </div>
  <div class="control-group">
    <label>Filter by Account Partner</label>
    <select id="filterPartner" onchange="applyFilters()">
      <option value="">All partners</option>
      {partner_options}
    </select>
  </div>
  <div class="control-group">
    <label>Sort by</label>
    <div class="sort-group">
      <button class="sort-btn" id="sortAccount" onclick="setSort('account')">Account</button>
      <button class="sort-btn active desc" id="sortPrice" onclick="setSort('price')">Price</button>
      <button class="sort-btn" id="sortPartner" onclick="setSort('partner')">Account Partner</button>
    </div>
  </div>
</div>

<div class="legend">
  <div class="legend-item"><div class="dot dot-yes"></div> Has services attach</div>
  <div class="legend-item"><div class="dot dot-no"></div> No services attach</div>
</div>
<div class="result-count" id="resultCount"></div>
<div class="wrapper">
<table>
<thead>
<tr>
  <th>Account</th>
  <th>License $</th>
  <th>Sales AE</th>
  <th>License Opportunity</th>
  <th>Close Date</th>
  <th class="svc-hdr">Services Opportunity</th>
  <th class="svc-hdr">Services $</th>
  <th class="svc-hdr">Account Partner</th>
  <th class="svc-hdr">Svc Close</th>
</tr>
</thead>
{rows}
</table>
</div>

<script>
var sortKey = 'price';
var sortDir = -1; // -1 = desc, 1 = asc

function getGroups() {{
  return Array.from(document.querySelectorAll('tbody.acct-group'));
}}

function applyFilters() {{
  var acctVal = document.getElementById('filterAccount').value.toLowerCase();
  var ptVal   = document.getElementById('filterPartner').value.toLowerCase();
  var visible = 0;
  getGroups().forEach(function(g) {{
    var acct = g.dataset.account.toLowerCase();
    var pt   = g.dataset.partner.toLowerCase();
    var show = (!acctVal || acct === acctVal) && (!ptVal || pt === ptVal);
    g.classList.toggle('hidden', !show);
    if (show) visible++;
  }});
  document.getElementById('resultCount').textContent =
    visible === getGroups().length ? '' : visible + ' of ' + getGroups().length + ' accounts shown';
}}

function setSort(key) {{
  if (sortKey === key) {{
    sortDir *= -1;
  }} else {{
    sortKey = key;
    sortDir = key === 'price' ? -1 : 1;
  }}
  ['account','price','partner'].forEach(function(k) {{
    var btn = document.getElementById('sort' + k.charAt(0).toUpperCase() + k.slice(1));
    btn.classList.remove('active','asc','desc');
  }});
  var active = document.getElementById('sort' + key.charAt(0).toUpperCase() + key.slice(1));
  active.classList.add('active', sortDir === 1 ? 'asc' : 'desc');
  doSort();
}}

function doSort() {{
  var table = document.querySelector('table');
  var groups = getGroups();
  groups.sort(function(a, b) {{
    var av, bv;
    if (sortKey === 'price') {{
      av = parseFloat(a.dataset.total) || 0;
      bv = parseFloat(b.dataset.total) || 0;
      return sortDir * (bv - av);  // always numeric desc for price default
    }} else if (sortKey === 'account') {{
      av = a.dataset.account.toLowerCase();
      bv = b.dataset.account.toLowerCase();
    }} else {{
      av = a.dataset.partner.toLowerCase();
      bv = b.dataset.partner.toLowerCase();
    }}
    if (av < bv) return -1 * sortDir;
    if (av > bv) return  1 * sortDir;
    return 0;
  }});
  groups.forEach(function(g) {{ table.appendChild(g); }});
}}

// fix price sort to honour direction
function doSort() {{
  var table = document.querySelector('table');
  var groups = getGroups();
  groups.sort(function(a, b) {{
    var av, bv;
    if (sortKey === 'price') {{
      av = parseFloat(a.dataset.total) || 0;
      bv = parseFloat(b.dataset.total) || 0;
      return sortDir === -1 ? (bv - av) : (av - bv);
    }} else if (sortKey === 'account') {{
      av = a.dataset.account.toLowerCase();
      bv = b.dataset.account.toLowerCase();
    }} else {{
      av = a.dataset.partner.toLowerCase();
      bv = b.dataset.partner.toLowerCase();
    }}
    if (av < bv) return -1 * sortDir;
    if (av > bv) return  1 * sortDir;
    return 0;
  }});
  groups.forEach(function(g) {{ table.appendChild(g); }});
}}

// init count
document.addEventListener('DOMContentLoaded', function() {{
  document.getElementById('resultCount').textContent = '';
}});
</script>
</body>
</html>""".format(
    acct_count=len(sorted_accounts),
    lic_count=len(license_rows),
    svc_count=len(services_rows),
    pipeline='{:,.0f}'.format(total_pipeline),
    accts_svc=accts_with_svc,
    accts_no_svc=len(sorted_accounts) - accts_with_svc,
    acct_options=acct_options,
    partner_options=partner_options,
    rows=rows_str
)

out_path = os.path.join(BASE, 'tmt_license_services_report.html')
with open(out_path, 'w', encoding='utf-8') as f:
    f.write(html_doc)

print("Written to {}".format(out_path))
print("Account groups: {}".format(len(tbodies_html)))
