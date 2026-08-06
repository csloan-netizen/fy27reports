import openpyxl
import html as html_module
import os

BASE = os.path.dirname(os.path.abspath(__file__))

wb = openpyxl.load_workbook(
    os.path.join(BASE, 'data', 'ap_pipegen.xlsx'),
    read_only=True, data_only=True
)

# ── Load Results (pivot: per-AP totals by quarter) ────────────────────────────
# Row 0: merged header (Q1 / Q2 / Q3 / Grand Total)
# Row 1: Geo, OU, Employee ID, Name, Bookings, Pipegen, Bookings, Pipegen, Bookings, Pipegen, Bookings, Pipegen
ws_res = wb['Results']
res_rows = list(ws_res.iter_rows(values_only=True))

results = {}  # emp_id -> dict
for r in res_rows[2:]:
    if not r[1] or 'TMT' not in str(r[1]):
        continue
    if not r[3]:  # skip blank name rows
        continue
    emp_id = str(r[2]) if r[2] is not None else ''
    results[emp_id] = {
        'name':          str(r[3]),
        'q1_bookings':   float(r[4])  if r[4]  is not None else 0,
        'q1_pipegen':    float(r[5])  if r[5]  is not None else 0,
        'q2_bookings':   float(r[6])  if r[6]  is not None else 0,
        'q2_pipegen':    float(r[7])  if r[7]  is not None else 0,
        'q3_bookings':   float(r[8])  if r[8]  is not None else 0,
        'q3_pipegen':    float(r[9])  if r[9]  is not None else 0,
        'tot_bookings':  float(r[10]) if r[10] is not None else 0,
        'tot_pipegen':   float(r[11]) if r[11] is not None else 0,
        'q3_open_pipe':  0,  # filled from Export below
    }

# ── Load Export (adds Q3 Open Pipe) ──────────────────────────────────────────
ws_exp = wb['Export']
exp_rows = list(ws_exp.iter_rows(values_only=True))
for r in exp_rows[1:]:
    if r[0] != 'Q3':
        continue
    if not r[7] or 'TMT' not in str(r[7]):
        continue
    emp_id = str(r[1]) if r[1] is not None else ''
    if emp_id in results:
        results[emp_id]['q3_open_pipe'] = float(r[9]) if r[9] is not None else 0

EXCLUDE = {
    'Pete Kilcommons', 'Robin Clayton', 'Elle Berry', 'Surabhi Kapoor',
    'Erin Porter', 'Sanjay Mittal', 'Taylor Hunt',
}
for emp_id in list(results.keys()):
    if results[emp_id]['name'] in EXCLUDE:
        del results[emp_id]

# Jason Buck has no rows in the source file; add him with zeros
ZERO = {k: 0 for k in ['q1_bookings','q1_pipegen','q2_bookings','q2_pipegen',
                        'q3_bookings','q3_pipegen','tot_bookings','tot_pipegen','q3_open_pipe']}
results['jason_buck'] = {'name': 'Jason Buck', **ZERO}

# ── Sort by Grand Total Pipegen descending ────────────────────────────────────
people = sorted(results.values(), key=lambda x: x['tot_pipegen'], reverse=True)

# ── Compute team totals ───────────────────────────────────────────────────────
def team_sum(field):
    return sum(p[field] for p in people)

totals = {f: team_sum(f) for f in [
    'q1_bookings','q1_pipegen','q2_bookings','q2_pipegen',
    'q3_bookings','q3_open_pipe','q3_pipegen','tot_bookings','tot_pipegen'
]}

# ── Helpers ───────────────────────────────────────────────────────────────────
def h(s):
    return html_module.escape(str(s) if s else '')

def fmt(v):
    if v == 0:
        return '<span class="zero">—</span>'
    return '${:,.0f}'.format(v)

def fmt_plain(v):
    if v == 0:
        return '—'
    return '${:,.0f}'.format(v)

def pct_bar(val, max_val, color):
    if max_val == 0:
        return ''
    pct = min(100, val / max_val * 100)
    return '<div class="bar-wrap"><div class="bar" style="width:{:.1f}%;background:{}"></div></div>'.format(pct, color)

max_tot_pipegen = people[0]['tot_pipegen'] if people else 1

# ── Build rows ────────────────────────────────────────────────────────────────
rows_html = []
for i, p in enumerate(people):
    bar = pct_bar(p['tot_pipegen'], max_tot_pipegen, '#0070d2')
    rows_html.append('''<tr class="data-row">
  <td class="name">{name}<div class="rank">#{rank}</div></td>
  <td class="money">{q1b}</td>
  <td class="money pipegen">{q1p}</td>
  <td class="money">{q2b}</td>
  <td class="money pipegen">{q2p}</td>
  <td class="money">{q3b}</td>
  <td class="money open">{q3op}</td>
  <td class="money pipegen">{q3p}</td>
  <td class="money tot-book">{tb}</td>
  <td class="money tot-pipe">{tp}<div class="bar-cell">{bar}</div></td>
</tr>'''.format(
        name=h(p['name']),
        rank=i+1,
        q1b=fmt(p['q1_bookings']),
        q1p=fmt(p['q1_pipegen']),
        q2b=fmt(p['q2_bookings']),
        q2p=fmt(p['q2_pipegen']),
        q3b=fmt(p['q3_bookings']),
        q3op=fmt(p['q3_open_pipe']),
        q3p=fmt(p['q3_pipegen']),
        tb=fmt(p['tot_bookings']),
        tp=fmt(p['tot_pipegen']),
        bar=bar,
    ))

# Team total row
rows_html.append('''<tr class="total-row">
  <td class="name">Team Total</td>
  <td class="money">{q1b}</td>
  <td class="money pipegen">{q1p}</td>
  <td class="money">{q2b}</td>
  <td class="money pipegen">{q2p}</td>
  <td class="money">{q3b}</td>
  <td class="money open">{q3op}</td>
  <td class="money pipegen">{q3p}</td>
  <td class="money tot-book">{tb}</td>
  <td class="money tot-pipe">{tp}</td>
</tr>'''.format(
    q1b=fmt(totals['q1_bookings']),
    q1p=fmt(totals['q1_pipegen']),
    q2b=fmt(totals['q2_bookings']),
    q2p=fmt(totals['q2_pipegen']),
    q3b=fmt(totals['q3_bookings']),
    q3op=fmt(totals['q3_open_pipe']),
    q3p=fmt(totals['q3_pipegen']),
    tb=fmt(totals['tot_bookings']),
    tp=fmt(totals['tot_pipegen']),
))

rows_str = '\n'.join(rows_html)

# ── Stat cards ────────────────────────────────────────────────────────────────
# Bookings pace: Q1+Q2+Q3 actual bookings vs pipegen
bookings_ytd = totals['tot_bookings']
pipegen_ytd  = totals['tot_pipegen']
attach_rate  = bookings_ytd / pipegen_ytd * 100 if pipegen_ytd else 0

stat_cards = [
    ('APs',             str(len(people)),                                          ''),
    ('FY27 Bookings',   '${:,.0f}'.format(bookings_ytd),                           'YTD Q1–Q3'),
    ('FY27 Pipegens',   '${:,.0f}'.format(pipegen_ytd),                            'YTD Q1–Q3'),
    ('Q3 Open Pipe',    '${:,.0f}'.format(totals['q3_open_pipe']),                  'Live pipeline'),
    ('Book/Pipe Ratio', '{:.1f}%'.format(attach_rate),                             'Bookings ÷ Pipegen'),
]

stats_html = '\n'.join(
    '<div class="stat"><div class="val">{}</div><div class="lbl">{}</div>{}</div>'.format(
        h(v), h(label),
        '<div class="sub">{}</div>'.format(h(sub)) if sub else ''
    )
    for label, v, sub in stat_cards
)

# ── Full HTML ─────────────────────────────────────────────────────────────────
html_doc = """<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<title>AMER TMT/CBS — AP PipeGen FY27</title>
<style>
* {{ box-sizing: border-box; margin: 0; padding: 0; }}
body {{ font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Arial, sans-serif; font-size: 12px; background: #f0f4f8; color: #1a1a2e; }}
.header {{ padding: 20px 24px 0; }}
h1 {{ font-size: 20px; color: #032D60; margin-bottom: 4px; }}
.subtitle {{ color: #555; font-size: 12px; margin-bottom: 14px; }}
.stats {{ display: flex; gap: 12px; padding: 0 24px 18px; flex-wrap: wrap; }}
.stat {{ background: white; border-radius: 6px; padding: 10px 18px; box-shadow: 0 1px 3px rgba(0,0,0,.1); min-width: 140px; }}
.stat .val {{ font-size: 22px; font-weight: 700; color: #032D60; }}
.stat .lbl {{ font-size: 10px; color: #666; margin-top: 2px; text-transform: uppercase; letter-spacing: .5px; }}
.stat .sub {{ font-size: 10px; color: #aaa; margin-top: 1px; }}
.wrapper {{ padding: 0 16px 40px; overflow-x: auto; }}
table {{ border-collapse: collapse; width: 100%; background: white; box-shadow: 0 1px 6px rgba(0,0,0,.1); border-radius: 8px; overflow: hidden; min-width: 900px; }}
thead tr.qhdr {{ background: #032D60; color: white; }}
thead tr.qhdr th {{ padding: 8px 10px; font-size: 10px; font-weight: 600; letter-spacing: .5px; text-align: center; }}
thead tr.qhdr th.q2hdr {{ background: #0b4f8a; }}
thead tr.qhdr th.q3hdr {{ background: #0070d2; }}
thead tr.qhdr th.tothdr {{ background: #00396b; }}
thead tr.qhdr th.namehdr {{ text-align: left; }}
thead tr.fhdr {{ background: #1a3a5c; color: #cde; }}
thead tr.fhdr th {{ padding: 7px 10px; font-size: 10px; font-weight: 500; letter-spacing: .3px; white-space: nowrap; }}
thead tr.fhdr th.pipegen {{ color: #88ccff; }}
thead tr.fhdr th.open {{ color: #aaddff; }}
td {{ padding: 7px 10px; vertical-align: middle; border-bottom: 1px solid #eef0f3; }}
tr.data-row:hover td {{ background: #f5f9ff; }}
tr.total-row td {{ background: #032D60; color: white; font-weight: 700; border-top: 2px solid #0070d2; }}
tr.total-row td.money {{ color: #cce5ff; }}
td.name {{ font-weight: 600; color: #032D60; width: 160px; border-right: 2px solid #dde3ed; white-space: nowrap; }}
td.name .rank {{ font-size: 10px; font-weight: 400; color: #aaa; display: inline; margin-left: 6px; }}
tr.total-row td.name {{ color: white; }}
td.money {{ text-align: right; white-space: nowrap; font-variant-numeric: tabular-nums; width: 100px; }}
td.money .zero {{ color: #ccc; font-style: normal; }}
td.pipegen {{ background: #f0f7ff; }}
td.open {{ background: #e8f4ff; }}
td.tot-book {{ background: #f5f0ff; font-weight: 600; }}
td.tot-pipe {{ background: #e8f0ff; font-weight: 700; color: #032D60; width: 130px; }}
tr.total-row td.tot-pipe {{ color: #88ccff; }}
.bar-cell {{ margin-top: 4px; }}
.bar-wrap {{ background: #dde8f5; border-radius: 2px; height: 4px; width: 100%; }}
.bar {{ height: 4px; border-radius: 2px; }}
</style>
</head>
<body>
<div class="header">
  <h1>AMER TMT/CBS — AP PipeGen FY27</h1>
  <p class="subtitle">Account Partner bookings &amp; pipeline generation &nbsp;·&nbsp; sorted by Grand Total Pipegen descending &nbsp;·&nbsp; as of 2026-08-06</p>
</div>
<div class="stats">
{stats}
</div>
<div class="wrapper">
<table>
<thead>
<tr class="qhdr">
  <th class="namehdr" rowspan="2">Account Partner</th>
  <th colspan="2">Q1</th>
  <th class="q2hdr" colspan="2">Q2</th>
  <th class="q3hdr" colspan="3">Q3</th>
  <th class="tothdr" colspan="2">Grand Total</th>
</tr>
<tr class="fhdr">
  <th>Bookings</th>
  <th class="pipegen">Pipegen</th>
  <th>Bookings</th>
  <th class="pipegen">Pipegen</th>
  <th>Bookings</th>
  <th class="open">Open Pipe</th>
  <th class="pipegen">Pipegen</th>
  <th>Bookings</th>
  <th class="pipegen">Pipegen</th>
</tr>
</thead>
<tbody>
{rows}
</tbody>
</table>
</div>
</body>
</html>""".format(
    stats=stats_html,
    rows=rows_str,
)

out_path = os.path.join(BASE, 'tmt_ap_pipegen_report.html')
with open(out_path, 'w', encoding='utf-8') as f:
    f.write(html_doc)

print("Written to {}".format(out_path))
print("APs: {}".format(len(people)))
print("Team pipegen: ${:,.0f}".format(totals['tot_pipegen']))
print("Team bookings: ${:,.0f}".format(totals['tot_bookings']))
